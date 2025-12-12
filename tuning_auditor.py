#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
tuning_auditor_robusto.py

Script robusto para aplicar TUNING / Evaluación de Experto
sobre la hoja 'Matriz_Riesgos' de un SGSI en Excel.

- Mantiene las macros VBA (keep_vba=True).
- Crea/actualiza la hoja 'Config_Tuning' con la matriz de evaluación del auditor.
- Crea columnas:
    * TUNING_AUDITOR   (nivel 1–5)
    * FACTOR_TUNING    (factor de ajuste 0,70–1,30)
    * RIESGO_TUNING    (riesgo inherente ajustado)
    * NIVEL_TUNING     (BAJO / MEDIO / ALTO / CRÍTICO)
- Colorea las celdas según el nivel de riesgo ajustado.

Autor: tú :)
"""

import sys
from typing import Optional, List

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import PatternFill, Font


# ---------------------------------------------------------------------------
# Parámetros principales
# ---------------------------------------------------------------------------

EXCEL_PATH_DEFAULT = "SGSI_COMPLETO_David_Nanculeo_v4.5_FINAL.xlsm"
HOJA_RIESGOS = "Matriz_Riesgos"
HOJA_TUNING = "Config_Tuning"
HEADER_ROW = 1

# Rango teórico del riesgo (por matriz prob * impacto)
RIESGO_MIN = 1
RIESGO_MAX = 25

# ---------------------------------------------------------------------------
# Utilidades generales
# ---------------------------------------------------------------------------

def log(msg: str) -> None:
    """Imprime mensajes de log simples."""
    print(f"[TUNING] {msg}")


def encontrar_columna_por_nombre(
    ws: Worksheet,
    posibles_nombres: List[str],
    header_row: int = HEADER_ROW
) -> Optional[int]:
    """
    Busca una columna por nombre en la fila de encabezados.

    :param ws: Hoja de Excel.
    :param posibles_nombres: Lista de nombres posibles (distintas variantes).
    :param header_row: Fila donde están los encabezados (por defecto 1).
    :return: Índice de columna (1-based) o None si no existe.
    """
    encabezados = {}
    for col in range(1, ws.max_column + 1):
        valor = ws.cell(row=header_row, column=col).value
        if valor is None:
            continue
        encabezados[str(valor).strip().upper()] = col

    for nombre in posibles_nombres:
        key = nombre.strip().upper()
        if key in encabezados:
            return encabezados[key]

    return None


def obtener_o_crear_columna(
    ws: Worksheet,
    titulo: str,
    header_row: int = HEADER_ROW
) -> int:
    """
    Devuelve el índice de columna para un título. Si no existe, lo crea
    en la siguiente columna libre.

    :param ws: Hoja de Excel.
    :param titulo: Título exacto de la columna a buscar/crear.
    :param header_row: Fila de encabezados.
    :return: Índice de columna (1-based).
    """
    # Buscar primero si ya existe
    col = encontrar_columna_por_nombre(ws, [titulo], header_row)
    if col is not None:
        return col

    # Crear en la siguiente columna disponible
    nueva_col = ws.max_column + 1
    ws.cell(row=header_row, column=nueva_col, value=titulo)
    # Formato básico de encabezado
    cell = ws.cell(row=header_row, column=nueva_col)
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid")
    return nueva_col


# ---------------------------------------------------------------------------
# Configuración de la matriz de Tuning del auditor
# ---------------------------------------------------------------------------

def configurar_hoja_tuning(wb) -> Worksheet:
    """
    Crea o actualiza la hoja Config_Tuning con la escala 1–5 y
    sus factores/interpretaciones.
    """
    if HOJA_TUNING in wb.sheetnames:
        ws_cfg = wb[HOJA_TUNING]
        # Limpiar contenido, pero mantener la hoja
        for row in ws_cfg.iter_rows():
            for cell in row:
                cell.value = None
    else:
        ws_cfg = wb.create_sheet(HOJA_TUNING)

    ws_cfg.title = HOJA_TUNING

    # Encabezados
    headers = ["Nivel (1-5)", "Etiqueta", "Factor_Tuning", "Interpretación", "Detalle técnico"]
    for col_idx, title in enumerate(headers, start=1):
        c = ws_cfg.cell(row=1, column=col_idx, value=title)
        c.font = Font(bold=True)
        c.fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")

    # Definición de escala: ±30% (0,15 por paso desde el neutro)
    niveles = [
        (1, "Muy Atenuante (-2)", 0.70,
         "Reduce fuertemente el riesgo calculado",
         "Controles compensatorios muy robustos que la fórmula estándar no refleja."),
        (2, "Atenuante (-1)", 0.85,
         "Reduce moderadamente el riesgo",
         "Controles adicionales efectivos, exposición menor a la esperada."),
        (3, "Neutro (0)", 1.00,
         "No ajusta el riesgo",
         "La fórmula Prob×Impacto coincide con la realidad observada."),
        (4, "Agravante (+1)", 1.15,
         "Aumenta moderadamente el riesgo",
         "Brechas relevantes, controles débiles o mala cultura de seguridad."),
        (5, "Muy Agravante (+2)", 1.30,
         "Aumenta fuertemente el riesgo",
         "Hallazgo crítico o vulnerabilidad explotable activa (suicidio digital)."),
    ]

    for row_idx, nivel in enumerate(niveles, start=2):
        ws_cfg.cell(row=row_idx, column=1, value=nivel[0])
        ws_cfg.cell(row=row_idx, column=2, value=nivel[1])
        ws_cfg.cell(row=row_idx, column=3, value=nivel[2])
        ws_cfg.cell(row=row_idx, column=4, value=nivel[3])
        ws_cfg.cell(row=row_idx, column=5, value=nivel[4])

    # Ajuste simple de ancho de columnas
    widths = [14, 24, 15, 40, 60]
    for col_idx, w in enumerate(widths, start=1):
        ws_cfg.column_dimensions[chr(ord('A') + col_idx - 1)].width = w

    log("Hoja 'Config_Tuning' creada/actualizada.")
    return ws_cfg


def obtener_factor_tuning(ws_cfg: Worksheet, nivel_tuning: int) -> float:
    """
    Busca el factor de Tuning para un nivel dado en la hoja Config_Tuning.

    :param ws_cfg: Hoja de configuración.
    :param nivel_tuning: Nivel 1–5.
    :return: Factor de ajuste (float). Devuelve 1.0 si no encuentra nada.
    """
    for row in range(2, ws_cfg.max_row + 1):
        nivel = ws_cfg.cell(row=row, column=1).value
        factor = ws_cfg.cell(row=row, column=3).value
        if nivel == nivel_tuning and isinstance(factor, (int, float)):
            return float(factor)
    return 1.0


# ---------------------------------------------------------------------------
# Lógica de riesgo y estilos
# ---------------------------------------------------------------------------

def clasificar_nivel_riesgo(riesgo: float) -> str:
    """
    Devuelve el nivel de riesgo (BAJO, MEDIO, ALTO, CRÍTICO)
    según el valor numérico.
    """
    if riesgo >= 15:
        return "CRÍTICO"
    elif riesgo >= 10:
        return "ALTO"
    elif riesgo >= 5:
        return "MEDIO"
    else:
        return "BAJO"


def aplicar_formato_riesgo(ws: Worksheet, fila: int, col_riesgo: int, col_nivel: int, riesgo: float) -> None:
    """
    Aplica formato de color a las celdas de riesgo y nivel ajustado
    según la escala de criticidad.
    """
    nivel = clasificar_nivel_riesgo(riesgo)

    if nivel == "CRÍTICO":
        fill_color = "C00000"  # Rojo oscuro
        font_color = "FFFFFF"
    elif nivel == "ALTO":
        fill_color = "FF6F00"  # Naranjo fuerte
        font_color = "FFFFFF"
    elif nivel == "MEDIO":
        fill_color = "FFC000"  # Amarillo
        font_color = "000000"
    else:  # BAJO
        fill_color = "92D050"  # Verde
        font_color = "000000"

    riesgo_cell = ws.cell(row=fila, column=col_riesgo)
    nivel_cell = ws.cell(row=fila, column=col_nivel)

    for c in (riesgo_cell, nivel_cell):
        c.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        c.font = Font(color=font_color, bold=True)


# ---------------------------------------------------------------------------
# Proceso principal de Tuning
# ---------------------------------------------------------------------------

def aplicar_tuning_auditor(excel_path: str = EXCEL_PATH_DEFAULT) -> None:
    """
    Aplica el Tuning del auditor sobre la hoja Matriz_Riesgos de un archivo Excel.

    - Busca la columna de 'Riesgo Inherente' por nombre.
    - Crea las columnas de TUNING y riesgo ajustado si no existen.
    - Usa la hoja Config_Tuning para traducir nivel 1–5 en factor.
    - Sobrescribe/crea un archivo nuevo con sufijo '_TUNING'.
    """
    log(f"Cargando archivo: {excel_path}")

    # Cargar libro manteniendo macros
    try:
        wb = load_workbook(excel_path, keep_vba=True)
    except FileNotFoundError:
        log(f"ERROR: No se encontró el archivo '{excel_path}'.")
        return
    except Exception as e:
        log(f"ERROR al abrir el libro: {e}")
        return

    # Verificar hoja de riesgos
    if HOJA_RIESGOS not in wb.sheetnames:
        log(f"ERROR: No se encontró la hoja '{HOJA_RIESGOS}' en el libro.")
        return

    ws_riesgos = wb[HOJA_RIESGOS]
    log(f"Hoja de riesgos detectada: '{HOJA_RIESGOS}'.")

    # Crear/actualizar Config_Tuning
    ws_cfg = configurar_hoja_tuning(wb)

    # Detectar columna de Riesgo Inherente
    col_riesgo_inh = encontrar_columna_por_nombre(
        ws_riesgos,
        posibles_nombres=["Riesgo Inherente", "RIESGO_INHERENTE", "RIESGO_INH", "Riesgo"]
    )
    if col_riesgo_inh is None:
        log("ADVERTENCIA: No se encontró la columna de 'Riesgo Inherente' por nombre.")
        log("Se asumirá la columna G (7) como riesgo inherente. Ajusta esto si tu estructura es distinta.")
        col_riesgo_inh = 7  # fallback duro, como en tu macro

    log(f"Columna de Riesgo Inherente: índice {col_riesgo_inh}.")

    # Columnas nuevas / existentes para Tuning
    col_tuning = obtener_o_crear_columna(ws_riesgos, "TUNING_AUDITOR")
    col_factor = obtener_o_crear_columna(ws_riesgos, "FACTOR_TUNING")
    col_riesgo_tun = obtener_o_crear_columna(ws_riesgos, "RIESGO_TUNING")
    col_nivel_tun = obtener_o_crear_columna(ws_riesgos, "NIVEL_TUNING")

    log(f"Columnas Tuning -> TUNING_AUDITOR={col_tuning}, FACTOR_TUNING={col_factor}, "
        f"RIESGO_TUNING={col_riesgo_tun}, NIVEL_TUNING={col_nivel_tun}.")

    # Opcional: detectar columna de ID para saber hasta dónde leer
    col_id = encontrar_columna_por_nombre(
        ws_riesgos,
        posibles_nombres=["ID", "ID_RIESGO", "Id Riesgo", "ID Riesgo"]
    )
    if col_id is None:
        col_id = 1  # usamos primera columna como referencia de fila "con datos"

    total_riesgos = 0
    filas_saltadas = 0

    # Recorrer filas de datos
    ultima_fila = ws_riesgos.max_row
    log(f"Procesando filas desde {HEADER_ROW + 1} hasta {ultima_fila}...")

    for fila in range(HEADER_ROW + 1, ultima_fila + 1):
        id_val = ws_riesgos.cell(row=fila, column=col_id).value
        if id_val is None or str(id_val).strip() == "":
            # Fila vacía / sin riesgo definido
            filas_saltadas += 1
            continue

        riesgo_inh = ws_riesgos.cell(row=fila, column=col_riesgo_inh).value

        # Si el riesgo inherente está vacío o no es numérico, saltamos la fila
        if riesgo_inh is None or not isinstance(riesgo_inh, (int, float)):
            filas_saltadas += 1
            continue

        # Leer nivel de Tuning (1–5). Si está vacío, asumimos neutro (3).
        tuning_val = ws_riesgos.cell(row=fila, column=col_tuning).value
        if tuning_val is None or tuning_val == "":
            tuning_val = 3
        try:
            tuning_val = int(tuning_val)
        except ValueError:
            log(f"Fila {fila}: valor de TUNING_AUDITOR no numérico ('{tuning_val}'), se fuerza a 3.")
            tuning_val = 3

        # Limitar a rango 1–5
        tuning_val = max(1, min(5, tuning_val))
        ws_riesgos.cell(row=fila, column=col_tuning, value=tuning_val)

        # Buscar factor en Config_Tuning
        factor = obtener_factor_tuning(ws_cfg, tuning_val)

        # Calcular riesgo ajustado
        riesgo_ajustado = float(riesgo_inh) * factor

        # Regla de override determinista: si tuning=5 y hay riesgo > 0,
        # asumimos que puede subir al máximo teórico.
        if tuning_val == 5 and riesgo_inh > 0:
            riesgo_ajustado = max(riesgo_ajustado, RIESGO_MAX)

        # Limitar a rango [RIESGO_MIN, RIESGO_MAX]
        riesgo_ajustado = max(RIESGO_MIN, min(RIESGO_MAX, riesgo_ajustado))

        # Redondear al entero más cercano
        riesgo_ajustado = round(riesgo_ajustado)

        # Escribir factor y riesgo ajustado
        ws_riesgos.cell(row=fila, column=col_factor, value=factor)
        ws_riesgos.cell(row=fila, column=col_riesgo_tun, value=riesgo_ajustado)

        # Escribir nivel de riesgo y aplicar formato
        nivel_texto = clasificar_nivel_riesgo(riesgo_ajustado)
        ws_riesgos.cell(row=fila, column=col_nivel_tun, value=nivel_texto)
        aplicar_formato_riesgo(ws_riesgos, fila, col_riesgo_tun, col_nivel_tun, riesgo_ajustado)

        total_riesgos += 1

    log(f"Riesgos ajustados con Tuning: {total_riesgos}. Filas saltadas: {filas_saltadas}.")

    # Guardar en un archivo nuevo para no pisar el original
    if excel_path.lower().endswith(".xlsm"):
        output_path = excel_path[:-5] + "_TUNING.xlsm"
    else:
        output_path = excel_path.rsplit(".", 1)[0] + "_TUNING.xlsx"

    try:
        wb.save(output_path)
        log(f"Archivo guardado con Tuning en: {output_path}")
    except Exception as e:
        log(f"ERROR al guardar el archivo con Tuning: {e}")


# ---------------------------------------------------------------------------
# Punto de entrada
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    # Permitir pasar el archivo por línea de comandos:
    #   python tuning_auditor_robusto.py mi_archivo.xlsm
    if len(sys.argv) > 1:
        excel_path = sys.argv[1]
    else:
        excel_path = EXCEL_PATH_DEFAULT

    aplicar_tuning_auditor(excel_path)

    
#EJECUTAR: python tuning_auditor.py
#ejectuar con archivo personalizado:
# python tuning_auditor.py ruta/a/mi_archivo_sgsi.xlsm