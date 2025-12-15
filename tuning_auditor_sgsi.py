#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
tuning_auditor_sgsi.py

Algoritmo de Tuning / Evaluación de Experto para el SGSI.

Alineado a:
- Libro: SGSI_COMPLETO_David_Nanculeo_v4.5_FINAL.xlsm (o similar)
- Hoja de riesgos: "Matriz_Riesgos"
- Fila de encabezados: 3
- Columna de riesgo base: "Nivel Riesgo Inherente"
- Columna de ID: "ID"

Qué hace:
- Crea/actualiza la hoja Config_Tuning con la escala 1–5.
- Asegura las columnas en Matriz_Riesgos:
    * TUNING_AUDITOR
    * FACTOR_TUNING
    * RIESGO_TUNING
    * NIVEL_TUNING
- Ajusta el riesgo base usando el factor de Tuning.
- Guarda una copia del archivo con sufijo "_TUNING".

Uso:

    python tuning_auditor_sgsi.py "SGSI_COMPLETO_David_Nanculeo_v4.5_FINAL.xlsm"

Si no se pasa argumento, usa ese nombre por defecto.
"""

import sys
from typing import List, Optional

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import PatternFill, Font


# --------- CONFIGURACIÓN ADAPTADA A TU XLSM ---------
HOJA_RIESGOS = "Matriz_Riesgos"
HOJA_TUNING = "Config_Tuning"

# En tu archivo, la fila 3 contiene los encabezados (ID, ID_Activo, etc.)
HEADER_ROW = 3

# Rango esperado de riesgo (por ejemplo, 1–25)
RIESGO_MIN = 1
RIESGO_MAX = 25


# ----------------------------------------------------
# UTILIDADES BÁSICAS
# ----------------------------------------------------
def log(msg: str) -> None:
    print(f"[TUNING] {msg}")


def encontrar_columna_por_nombre(
    ws: Worksheet,
    posibles_nombres: List[str],
    header_row: int = HEADER_ROW,
) -> Optional[int]:
    """
    Busca la columna cuyo encabezado coincida con alguno de los nombres dados.
    Devuelve índice 1-based o None si no existe.
    """
    last_col = ws.max_column
    headers = {}

    for col in range(1, last_col + 1):
        val = ws.cell(row=header_row, column=col).value
        if val is None:
            continue
        headers[str(val).strip().upper()] = col

    for nombre in posibles_nombres:
        key = nombre.strip().upper()
        if key in headers:
            return headers[key]

    return None


def obtener_o_crear_columna(
    ws: Worksheet,
    titulo: str,
    header_row: int = HEADER_ROW,
) -> int:
    """
    Devuelve el índice de la columna con el encabezado "titulo".
    Si no existe, la crea al final.
    """
    col = encontrar_columna_por_nombre(ws, [titulo], header_row)
    if col is not None:
        return col

    new_col = ws.max_column + 1
    cell = ws.cell(row=header_row, column=new_col)
    cell.value = titulo
    cell.font = Font(bold=True)
    cell.fill = PatternFill(
        start_color="E2F0D9",
        end_color="E2F0D9",
        fill_type="solid",
    )
    log(f"Creada columna '{titulo}' en la posición {new_col}.")
    return new_col


# ----------------------------------------------------
# CONFIG_TUNING
# ----------------------------------------------------
def configurar_hoja_tuning(wb) -> Worksheet:
    """
    Crea o limpia la hoja Config_Tuning y escribe la tabla de
    niveles de Tuning (1–5) con sus factores.
    """
    if HOJA_TUNING in wb.sheetnames:
        ws_cfg = wb[HOJA_TUNING]
        for row in ws_cfg.iter_rows():
            for c in row:
                c.value = None
    else:
        ws_cfg = wb.create_sheet(HOJA_TUNING)

    ws_cfg.title = HOJA_TUNING

    headers = ["Nivel (1-5)", "Etiqueta", "Factor_Tuning", "Interpretación", "Detalle técnico"]
    for i, title in enumerate(headers, start=1):
        c = ws_cfg.cell(row=1, column=i, value=title)
        c.font = Font(bold=True)
        c.fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")

    niveles = [
        (1, "Muy Atenuante (-2)", 0.70,
         "Reduce fuertemente el riesgo calculado",
         "Controles compensatorios muy robustos que la fórmula estándar no refleja."),
        (2, "Atenuante (-1)", 0.85,
         "Reduce moderadamente el riesgo",
         "Controles adicionales efectivos, exposición menor a la esperada."),
        (3, "Neutro (0)", 1.00,
         "No ajusta el riesgo",
         "La fórmula Prob×Impacto coincide con la realidad observada."),
        (4, "Agravante (+1)", 1.15,
         "Aumenta moderadamente el riesgo",
         "Brechas relevantes, controles débiles o mala cultura de seguridad."),
        (5, "Muy Agravante (+2)", 1.30,
         "Aumenta fuertemente el riesgo",
         "Hallazgo crítico o vulnerabilidad explotable activa."),
    ]

    for r, nivel in enumerate(niveles, start=2):
        ws_cfg.cell(row=r, column=1, value=nivel[0])
        ws_cfg.cell(row=r, column=2, value=nivel[1])
        ws_cfg.cell(row=r, column=3, value=nivel[2])
        ws_cfg.cell(row=r, column=4, value=nivel[3])
        ws_cfg.cell(row=r, column=5, value=nivel[4])

    # Anchos para que quede legible
    widths = [14, 26, 16, 42, 60]
    for i, w in enumerate(widths, start=1):
        col_letter = chr(ord("A") + i - 1)
        ws_cfg.column_dimensions[col_letter].width = w

    log("Hoja 'Config_Tuning' creada/actualizada.")
    return ws_cfg


def obtener_factor_tuning(ws_cfg: Worksheet, nivel_tuning: int) -> float:
    """
    Devuelve el factor (0.70, 0.85, 1.00, 1.15, 1.30) para el nivel dado (1–5).
    Si no lo encuentra, devuelve 1.0.
    """
    last_row = ws_cfg.max_row
    for row in range(2, last_row + 1):
        nivel = ws_cfg.cell(row=row, column=1).value
        factor = ws_cfg.cell(row=row, column=3).value
        if nivel == nivel_tuning and isinstance(factor, (int, float)):
            return float(factor)
    return 1.0


# ----------------------------------------------------
# LÓGICA DE CLASIFICACIÓN Y FORMATO
# ----------------------------------------------------
def clasificar_nivel_riesgo(riesgo: float) -> str:
    """
    Clasifica el riesgo numérico en BAJO / MEDIO / ALTO / CRÍTICO.
    Estos umbrales puedes ajustarlos si quieres.
    """
    if riesgo >= 15:
        return "CRÍTICO"
    elif riesgo >= 10:
        return "ALTO"
    elif riesgo >= 5:
        return "MEDIO"
    else:
        return "BAJO"


def aplicar_formato_riesgo(
    ws: Worksheet,
    fila: int,
    col_riesgo: int,
    col_nivel: int,
    riesgo: float,
) -> None:
    nivel = clasificar_nivel_riesgo(riesgo)

    if nivel == "CRÍTICO":
        fill_color = "C00000"
        font_color = "FFFFFF"
    elif nivel == "ALTO":
        fill_color = "FF6F00"
        font_color = "FFFFFF"
    elif nivel == "MEDIO":
        fill_color = "FFC000"
        font_color = "000000"
    else:  # BAJO
        fill_color = "92D050"
        font_color = "000000"

    c_riesgo = ws.cell(row=fila, column=col_riesgo)
    c_nivel = ws.cell(row=fila, column=col_nivel)

    for c in (c_riesgo, c_nivel):
        c.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        c.font = Font(color=font_color, bold=True)


# ----------------------------------------------------
# CÁLCULO DEL TUNING
# ----------------------------------------------------
def normalizar_tuning(valor) -> int:
    """
    Convierte cualquier valor a un entero entre 1 y 5.
    Si no es numérico, devuelve 3 (Neutro).
    """
    try:
        n = int(valor)
    except (TypeError, ValueError):
        return 3
    return max(1, min(5, n))


def calcular_riesgo_ajustado(riesgo_base: float, factor: float, nivel_tuning: int) -> int:
    """
    Aplica el factor de Tuning sobre el riesgo base y lo acota a [RIESGO_MIN, RIESGO_MAX].
    Incluye una regla opcional: si nivel=5 (Muy Agravante) y el riesgo base > 0,
    se puede elevar al máximo teórico.
    """
    r = riesgo_base * factor

    if nivel_tuning == 5 and riesgo_base > 0:
        if r < RIESGO_MAX:
            r = RIESGO_MAX

    if r < RIESGO_MIN:
        r = RIESGO_MIN
    if r > RIESGO_MAX:
        r = RIESGO_MAX

    return int(round(r))


# ----------------------------------------------------
# FUNCIÓN PRINCIPAL
# ----------------------------------------------------
def aplicar_tuning_auditor(excel_path: str) -> None:
    log(f"Cargando archivo: {excel_path}")

    try:
        wb = load_workbook(excel_path, keep_vba=True)
    except FileNotFoundError:
        log(f"ERROR: archivo no encontrado: {excel_path}")
        return
    except Exception as e:
        log(f"ERROR al abrir el libro: {e}")
        return

    if HOJA_RIESGOS not in wb.sheetnames:
        log(f"ERROR: no existe la hoja '{HOJA_RIESGOS}' en el libro.")
        return

    ws_riesgos = wb[HOJA_RIESGOS]
    ws_cfg = configurar_hoja_tuning(wb)

    # Columna de riesgo base ("Nivel Riesgo Inherente" en tu matriz)
    col_riesgo_inh = encontrar_columna_por_nombre(
        ws_riesgos,
        [
            "Nivel Riesgo Inherente",
            "NIVEL RIESGO INHERENTE",
            "Riesgo Inherente",
            "RIESGO INHERENTE",
            "Riesgo",
            "RIESGO",
        ],
        header_row=HEADER_ROW,
    )
    if col_riesgo_inh is None:
        log("Advertencia: no se encontró la columna de riesgo inherente; usando columna 13 por defecto.")
        col_riesgo_inh = 13

    # ID de riesgo ("ID" en tu matriz)
    col_id = encontrar_columna_por_nombre(
        ws_riesgos,
        ["ID", "ID_RIESGO", "Id Riesgo", "ID Riesgo"],
        header_row=HEADER_ROW,
    )
    if col_id is None:
        col_id = 1

    # Columnas de Tuning (se crean si no existen)
    col_tuning = obtener_o_crear_columna(ws_riesgos, "TUNING_AUDITOR", header_row=HEADER_ROW)
    col_factor = obtener_o_crear_columna(ws_riesgos, "FACTOR_TUNING", header_row=HEADER_ROW)
    col_riesgo_tun = obtener_o_crear_columna(ws_riesgos, "RIESGO_TUNING", header_row=HEADER_ROW)
    col_nivel_tun = obtener_o_crear_columna(ws_riesgos, "NIVEL_TUNING", header_row=HEADER_ROW)

    last_row = ws_riesgos.max_row
    log(f"Procesando filas de {HEADER_ROW + 1} a {last_row}...")

    filas_proc = 0
    filas_skip = 0

    for fila in range(HEADER_ROW + 1, last_row + 1):
        id_val = ws_riesgos.cell(row=fila, column=col_id).value
        if id_val is None or str(id_val).strip() == "":
            filas_skip += 1
            continue

        riesgo_base = ws_riesgos.cell(row=fila, column=col_riesgo_inh).value
        if not isinstance(riesgo_base, (int, float)):
            filas_skip += 1
            continue

        tuning_val = normalizar_tuning(ws_riesgos.cell(row=fila, column=col_tuning).value)
        ws_riesgos.cell(row=fila, column=col_tuning).value = tuning_val

        factor = obtener_factor_tuning(ws_cfg, tuning_val)
        riesgo_adj = calcular_riesgo_ajustado(float(riesgo_base), factor, tuning_val)

        ws_riesgos.cell(row=fila, column=col_factor).value = factor
        ws_riesgos.cell(row=fila, column=col_riesgo_tun).value = riesgo_adj
        ws_riesgos.cell(row=fila, column=col_nivel_tun).value = clasificar_nivel_riesgo(riesgo_adj)

        aplicar_formato_riesgo(ws_riesgos, fila, col_riesgo_tun, col_nivel_tun, riesgo_adj)

        filas_proc += 1

    log(f"Filas procesadas: {filas_proc}, filas saltadas: {filas_skip}")

    # Guardar copia con sufijo _TUNING
    if excel_path.lower().endswith(".xlsm"):
        out_path = excel_path[:-5] + "_TUNING.xlsm"
    else:
        out_path = excel_path.rsplit(".", 1)[0] + "_TUNING.xlsx"

    try:
        wb.save(out_path)
        log(f"Archivo guardado: {out_path}")
    except Exception as e:
        log(f"ERROR al guardar el archivo: {e}")


# ----------------------------------------------------
# ENTRYPOINT
# ----------------------------------------------------
if __name__ == "__main__":
    if len(sys.argv) > 1:
        ruta_excel = sys.argv[1]
    else:
        ruta_excel = "SGSI_COMPLETO_David_Nanculeo_v4.5_FINAL.xlsm"

    aplicar_tuning_auditor(ruta_excel)
